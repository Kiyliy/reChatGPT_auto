import requests
import openpyxl
from queue import Queue
import json

###config
config_work_book_name = "_GPT账号管理Tokens1.xlsx"  #excel name
config_work_sheets_n = 0 #第一个工作表, default 0
config_fkrow = 6 #fk in the excel ( 0 is the start)
config_pk_cnt = 50 #多少个fk生成一个pk , default 50
config_write_book_name ="pk_tokens.xlsx"

#pk为空时, 生成pk, 否则更新pk
#多个fk使用\n分开
def pk_tokens(pool_token: str | None ,
              share_tokens:str
                  )-> json or None:
    url = "https://ai.fakeopen.com/pool/update"

    headers = {
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Origin": "https://ai.fakeopen.com",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36 Edg/115.0.1901.200",
        "X-Requested-With": "XMLHttpRequest"
    }

    data = {
        "share_tokens": share_tokens,
        "pool_token": pool_token
    }

    response = requests.post(url, headers=headers, data=data)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"请求失败，状态码 {response.status_code}")
        return None
    

def GetXlsxData(config_work_book_name,config_work_sheets_n,max_col=2):
    """
    参数: 工作簿名: str
        工作表序号: int
        最大列数
    返回值:队列
    """
    # 打开Excel文件
    workbook = openpyxl.load_workbook(config_work_book_name,read_only=True)
    # 获取工作表（可以通过名称或索引获取）
    worksheet = workbook.worksheets[config_work_sheets_n] # 或者使用 workbook.worksheets[0]
    rows_queue = Queue()
    #遍历工作表的行
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=max_col,values_only=True):
        rows_queue.put(row)
    return rows_queue


#生成pks
def pk_tokens_main():
    fk_tokens= GetXlsxData(config_work_book_name,config_work_sheets_n,max_col=7)
    pk_tokens_list = []
    fk_list = []
    fk_cnt = 0
    while not fk_tokens.empty():
        fk_list.append(fk_tokens.get()[config_fkrow])
        fk_cnt += 1
        if fk_cnt == config_pk_cnt:
            fk_cnt = 0
            fk_str = "\n".join(fk_list)
            pk_tokens_list.append(pk_tokens(pool_token=None,share_tokens=fk_str))
            fk_list = []
    if fk_cnt != 0:
        fk_str = "\n".join(fk_list)
        pk_tokens_list.append(pk_tokens(pool_token=None,share_tokens=fk_str))
    return pk_tokens_list
    
def main():
    pk_tokens_list = pk_tokens_main()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for i in range(len(pk_tokens_list)):
        worksheet.cell(row=i+1,column=1,value=pk_tokens_list[i]["pool_token"])
    workbook.save(config_write_book_name)
    print("pk_tokens.xlsx 已生成")

   
        

if __name__ == "__main__":
    main()