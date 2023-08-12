import openpyxl
import json
import config
from queue import Queue

def GetXlsxData(work_book_name,work_sheets_n,max_col=2):
    """
    参数: 工作簿名: str
        工作表序号: int
        最大列数
    返回值:队列
    """
    # 打开Excel文件
    workbook = openpyxl.load_workbook(work_book_name,read_only=True)
    # 获取工作表（可以通过名称或索引获取）
    worksheet = workbook.worksheets[work_sheets_n] # 或者使用 workbook.worksheets[0]
    rows_queue = Queue()
    #遍历工作表的行
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=max_col,values_only=True):
        rows_queue.put(row)
    return rows_queue


def WriteAccessTokenToFile(username: str, 
                           password: str,
                           response :json,
                           AccessTokensWorksheet : openpyxl.worksheet.worksheet.Worksheet,
                           fakeresponse :json ,
                           errorWorksheet
                           ):

        if isinstance(response,str):
            errorWorksheet.append([username,password,"writer_err_as_e: "+response])
            return 
        elif response.status_code == 200:
            rsp = response.json()
            fkrsp = fakeresponse.json()
            #成功
            AccessTokensWorksheet.append([username, password, rsp["access_token"],\
                                        rsp["expires_in"],rsp["id_token"],rsp["refresh_token"],fkrsp["token_key"],fkrsp["unique_name"]])
        elif response.status_code != 200:
            try:
                errorWorksheet.append([username,password,"status_code_err: "+response.text])
            except:
                try:
                    errorWorksheet.append([username,password,"测试"])
                except:
                    print("写入失败!")
           

        
